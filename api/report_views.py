"""
api/report_views.py
=============================================================================
Roll-date and full-schedule reports for the logged-in supervisor's group.

Top Reports tab:
  - default as_of = active operating date
  - shows generated work from selected month start up to roll date

Bottom Full · Report tab:
  - ?as_of=all
  - shows the whole generated schedule for the selected month

The same metrics are returned for both scopes: jobs, hours, avg/day,
utilization, route distance, travel time, and callback SLA where applicable.
=============================================================================
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.db.models import Min, Max
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from api.models import Schedule, OperationType
from api.active_day import get_active_date
from api.services.maps.route_geometry import build_route_geometry


SHIFT_HOURS = 8.0


def _as_of(request):
    """Upper date bound the frontend is allowed to see.

    Default: operating-clock day.
    ?as_of=YYYY-MM-DD: explicit clamp.
    ?as_of=all: no clamp, used by Full · Report.
    """
    raw = (request.query_params.get("as_of")
           if hasattr(request, "query_params") else None)
    if raw:
        if raw.lower() == "all":
            return None
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return get_active_date()


def _supervised_group(request):
    return getattr(request.user, "supervised_group", None)


def _group_schedules(group, year=None, month=None, as_of=None):
    """All schedules for a group, optionally filtered to a month and clamped."""
    qs = (Schedule.objects
          .filter(technician__group=group,
                  technician__is_active_employee=True,
                  start_time__isnull=False)
          .select_related("technician", "task", "task__unit", "task__task_type"))
    if year and month:
        qs = qs.filter(start_time__year=year, start_time__month=month)
    if as_of is not None:
        qs = qs.filter(start_time__date__lte=as_of)
    return qs.order_by("technician__full_name", "start_time", "sequence_order")


def _minutes(s):
    if s.start_time and s.end_time:
        return max(0, int((s.end_time - s.start_time).total_seconds() // 60))
    return 0


def _km(value):
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _sla_met(s):
    task = s.task
    if not task or not task.task_type:
        return None
    if task.task_type.operation_type != OperationType.CALLBACK:
        return None
    if not task.latest_finish or not s.end_time:
        return None
    return s.end_time <= task.latest_finish


def _sort_technicians(rows, sort, order):
    reverse = (order or "asc").lower() == "desc"
    sort = sort or "name"

    def key(row):
        if sort == "hours":
            return row.get("total_hours", 0)
        if sort in ("jobs", "buildings"):
            return row.get("jobs", row.get("buildings_visited", 0))
        if sort == "days":
            return row.get("days_worked", 0)
        if sort == "utilization":
            return row.get("utilization_pct", 0)
        if sort == "route_km":
            return row.get("route_km", 0)
        if sort == "sla":
            return row.get("sla_pct", 0)
        return row.get("name", "")

    rows.sort(key=key, reverse=reverse)
    return rows


def _build_report(group, year, month, as_of=None, search="", sort="name", order="asc"):
    """Return per-technician aggregated report for the month/scope."""
    scheds = list(_group_schedules(group, year, month, as_of))

    per_tech = defaultdict(lambda: defaultdict(list))
    tech_info = {}

    for s in scheds:
        t = s.technician
        if search and search.lower() not in t.full_name.lower():
            continue

        task = s.task
        task_type = task.task_type if task else None
        unit = task.unit if task else None
        op = task_type.operation_type if task_type else ""
        priority = task.priority if task else None
        travel_min = s.travel_time_min or 0
        route_km = _km(s.travel_distance_km)
        sla = _sla_met(s)

        tech_info[t.id] = {
            "name": t.full_name,
            "role": t.tech_role,
            "specialty": t.specialty,
            "origin": (
                float(t.current_latitude),
                float(t.current_longitude),
            ) if t.current_latitude is not None and t.current_longitude is not None else None,
        }
        day = s.start_time.date().isoformat()
        per_tech[t.id][day].append({
            "building": unit.unit_name if unit else "Unknown unit",
            "unit_code": unit.unit_code if unit else None,
            "latitude": float(unit.latitude) if unit and unit.latitude is not None else None,
            "longitude": float(unit.longitude) if unit and unit.longitude is not None else None,
            "operation_type": op,
            "maintenance_type": (task_type.maintenance_type if task_type else None),
            "priority": priority,
            "start": s.start_time.strftime("%H:%M"),
            "end": s.end_time.strftime("%H:%M") if s.end_time else None,
            "minutes": _minutes(s),
            "travel_min": travel_min,
            "route_km": round(route_km, 2),
            "sla_met": sla,
            "task_no": task.task_no if task else None,
        })

    technicians = []
    summary_units = set()
    summary_dates = set()
    summary_jobs = 0
    summary_work_min = 0
    summary_travel_min = 0
    summary_route_km = 0.0
    summary_aa = 0
    summary_b = 0
    summary_sla_met = 0
    summary_sla_total = 0
    operation_types = set()

    for tid, days in per_tech.items():
        total_min = 0
        total_travel_min = 0
        total_route_km = 0.0
        total_visits = 0
        aa_count = 0
        b_count = 0
        sla_met_count = 0
        sla_total = 0
        day_list = []

        for day in sorted(days.keys()):
            visits = days[day]
            day_min = sum(v["minutes"] for v in visits)

            # Route KM / travel time are intentionally read from the cached
            # Google route geometry, not from Schedule.travel_distance_km.
            # The live map/mobile route cache is created by:
            #   python manage.py precache_google_routes YYYY-MM-DD YYYY-MM-DD
            # Report views must be billing-safe, so allow_google_call=False:
            # this reads CACHE if present and never calls Google while a report
            # page is opened/refreshed.
            info = tech_info.get(tid, {})
            origin = info.get("origin")
            route_points = [
                {"lat": v["latitude"], "lng": v["longitude"]}
                for v in visits
                if v.get("latitude") is not None and v.get("longitude") is not None
            ]
            day_route_source = None
            day_route = 0.0
            day_travel = sum(v["travel_min"] for v in visits)
            if route_points:
                if origin is None:
                    origin = (route_points[0]["lat"], route_points[0]["lng"])
                geom = build_route_geometry(origin, route_points, allow_google_call=False)
                day_route_source = geom.get("source")
                # Use only real Google/cache distance for management reporting.
                # If cache is missing, route_km stays 0 and the user should run
                # precache_google_routes for that report date range.
                if day_route_source in {"CACHE", "GOOGLE_ROADS"}:
                    day_route = float(geom.get("distance_km") or 0)
                    day_travel = int(geom.get("duration_min") or day_travel or 0)

            # Give each visit an equal share so the detail/export sheet is not
            # filled with zeros while the daily and technician total remains exact.
            route_share = round(day_route / len(visits), 2) if visits and day_route else 0.0
            travel_share = int(round(day_travel / len(visits))) if visits and day_travel else 0
            for v in visits:
                v["route_km"] = route_share
                v["travel_min"] = travel_share
                v["route_source"] = day_route_source

            total_min += day_min
            total_travel_min += day_travel
            total_route_km += day_route
            total_visits += len(visits)
            summary_dates.add(day)
            summary_jobs += len(visits)
            summary_work_min += day_min
            summary_travel_min += day_travel
            summary_route_km += day_route

            for v in visits:
                if v.get("unit_code"):
                    summary_units.add(str(v["unit_code"]))
                if v.get("operation_type"):
                    operation_types.add(str(v["operation_type"]))
                if v.get("priority") == "AA":
                    aa_count += 1
                    summary_aa += 1
                elif v.get("priority") == "B":
                    b_count += 1
                    summary_b += 1
                if v.get("sla_met") is not None:
                    sla_total += 1
                    summary_sla_total += 1
                    if v.get("sla_met"):
                        sla_met_count += 1
                        summary_sla_met += 1

            window = f'{visits[0]["start"]}–{visits[-1]["end"] or visits[-1]["start"]}'
            day_list.append({
                "date": day,
                "visits": visits,
                "buildings": len(visits),
                "work_minutes": day_min,
                "travel_minutes": day_travel,
                "route_km": round(day_route, 1),
                "route_source": day_route_source,
                "window": window,
            })

        days_worked = len(days)
        total_hours = round(total_min / 60, 1)
        avg_day_hours = round((total_min / 60) / days_worked, 1) if days_worked else 0.0
        utilization_pct = round((avg_day_hours / SHIFT_HOURS) * 100, 1) if days_worked else 0.0
        sla_pct = round((sla_met_count / sla_total) * 100, 1) if sla_total else 0.0
        info = tech_info[tid]
        technicians.append({
            "id": tid,
            "name": info["name"],
            "role": info["role"],
            "specialty": info["specialty"],
            "days_worked": days_worked,
            "buildings_visited": total_visits,  # kept for old frontend compatibility
            "jobs": total_visits,
            "total_hours": total_hours,
            "avg_day_hours": avg_day_hours,
            "utilization_pct": utilization_pct,
            "travel_minutes": total_travel_min,
            "travel_hours": round(total_travel_min / 60, 1),
            "route_km": round(total_route_km, 1),
            "aa_count": aa_count,
            "b_count": b_count,
            "sla_met": sla_met_count,
            "sla_total": sla_total,
            "sla_pct": sla_pct,
            "days": day_list,
        })

    _sort_technicians(technicians, sort, order)

    tech_days = sum(t["days_worked"] for t in technicians)
    avg_day_hours = round((summary_work_min / 60) / tech_days, 1) if tech_days else 0.0
    utilization_pct = round((avg_day_hours / SHIFT_HOURS) * 100, 1) if tech_days else 0.0
    summary_sla_pct = round((summary_sla_met / summary_sla_total) * 100, 1) if summary_sla_total else 0.0

    summary = {
        "technician_count": len(technicians),
        "jobs": summary_jobs,
        "units": len(summary_units),
        "hours": round(summary_work_min / 60, 1),
        "avg_day_hours": avg_day_hours,
        "utilization_pct": utilization_pct,
        "travel_minutes": summary_travel_min,
        "travel_hours": round(summary_travel_min / 60, 1),
        "route_km": round(summary_route_km, 1),
        "scheduled_days": len(summary_dates),
        "aa_count": summary_aa,
        "b_count": summary_b,
        "sla_met": summary_sla_met,
        "sla_total": summary_sla_total,
        "sla_pct": summary_sla_pct,
        "operation_types": sorted(operation_types),
    }
    return technicians, summary


class ReportMonthsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group = _supervised_group(request)
        if group is None:
            return Response({"error": "Not a supervisor of any group."}, status=403)

        rng = (_group_schedules(group, as_of=_as_of(request))
               .aggregate(lo=Min("start_time"), hi=Max("start_time")))
        months = []
        if rng["lo"] and rng["hi"]:
            y, m = rng["lo"].year, rng["lo"].month
            end_y, end_m = rng["hi"].year, rng["hi"].month
            while (y, m) <= (end_y, end_m):
                months.append({"year": y, "month": m,
                               "label": date(y, m, 1).strftime("%B %Y")})
                m += 1
                if m > 12:
                    m = 1
                    y += 1
        return Response({"months": months})


class MonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group = _supervised_group(request)
        if group is None:
            return Response({"error": "Not a supervisor of any group."}, status=403)
        try:
            year = int(request.query_params.get("year"))
            month = int(request.query_params.get("month"))
        except (TypeError, ValueError):
            return Response({"error": "year and month required."}, status=400)

        technicians, summary = _build_report(
            group, year, month, _as_of(request),
            search=(request.query_params.get("search") or "").strip(),
            sort=(request.query_params.get("sort") or "name"),
            order=(request.query_params.get("order") or "asc"),
        )
        return Response({
            "group": group.name,
            "year": year,
            "month": month,
            "label": date(year, month, 1).strftime("%B %Y"),
            "technician_count": len(technicians),
            "summary": summary,
            "technicians": technicians,
        })


class MonthlyReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group = _supervised_group(request)
        if group is None:
            return Response({"error": "Not a supervisor of any group."}, status=403)
        try:
            year = int(request.query_params.get("year"))
            month = int(request.query_params.get("month"))
        except (TypeError, ValueError):
            return Response({"error": "year and month required."}, status=400)

        technicians, summary = _build_report(
            group, year, month, _as_of(request),
            search=(request.query_params.get("search") or "").strip(),
            sort=(request.query_params.get("sort") or "name"),
            order=(request.query_params.get("order") or "asc"),
        )
        label = date(year, month, 1).strftime("%B_%Y")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return Response(
                {"error": "openpyxl not installed. Run: pip install openpyxl"},
                status=500)

        wb = openpyxl.Workbook()
        head_fill = PatternFill("solid", fgColor="1F4E78")
        head_font = Font(bold=True, color="FFFFFF")

        ws = wb.active
        ws.title = "Summary"
        headers = [
            "Technician", "Role", "Specialty", "Days", "Jobs",
            "Hours", "Avg/Day", "Utilization %", "Route KM",
            "Travel Hours", "AA", "B", "SLA %"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = head_fill
            cell.font = head_font
        for r, t in enumerate(technicians, 2):
            ws.cell(row=r, column=1, value=t["name"])
            ws.cell(row=r, column=2, value=t["role"])
            ws.cell(row=r, column=3, value=t["specialty"])
            ws.cell(row=r, column=4, value=t["days_worked"])
            ws.cell(row=r, column=5, value=t["jobs"])
            ws.cell(row=r, column=6, value=t["total_hours"])
            ws.cell(row=r, column=7, value=t["avg_day_hours"])
            ws.cell(row=r, column=8, value=t["utilization_pct"])
            ws.cell(row=r, column=9, value=t["route_km"])
            ws.cell(row=r, column=10, value=t["travel_hours"])
            ws.cell(row=r, column=11, value=t["aa_count"])
            ws.cell(row=r, column=12, value=t["b_count"])
            ws.cell(row=r, column=13, value=t["sla_pct"] if t["sla_total"] else None)
        for col, w in zip("ABCDEFGHIJKLM", [26, 14, 14, 10, 10, 10, 10, 14, 12, 14, 8, 8, 10]):
            ws.column_dimensions[col].width = w

        ws2 = wb.create_sheet("Detail")
        dheaders = [
            "Technician", "Date", "Start", "End", "Unit/Building", "Operation",
            "Type/Priority", "Minutes", "Travel (min)", "Route KM", "SLA Met"
        ]
        for col, h in enumerate(dheaders, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = head_fill
            cell.font = head_font
        row = 2
        for t in technicians:
            for day in t["days"]:
                for v in day["visits"]:
                    ws2.cell(row=row, column=1, value=t["name"])
                    ws2.cell(row=row, column=2, value=day["date"])
                    ws2.cell(row=row, column=3, value=v["start"])
                    ws2.cell(row=row, column=4, value=v["end"])
                    ws2.cell(row=row, column=5, value=v["building"])
                    ws2.cell(row=row, column=6, value=v["operation_type"])
                    ws2.cell(row=row, column=7, value=v["priority"] or v["maintenance_type"])
                    ws2.cell(row=row, column=8, value=v["minutes"])
                    ws2.cell(row=row, column=9, value=v["travel_min"])
                    ws2.cell(row=row, column=10, value=v["route_km"])
                    ws2.cell(row=row, column=11, value=("YES" if v["sla_met"] else "NO") if v["sla_met"] is not None else "")
                    row += 1
        for col, w in zip("ABCDEFGHIJK", [24, 12, 8, 8, 34, 14, 12, 10, 12, 10, 10]):
            ws2.column_dimensions[col].width = w

        ws3 = wb.create_sheet("KPI")
        for r, (k, v) in enumerate(summary.items(), 1):
            ws3.cell(row=r, column=1, value=k)
            ws3.cell(row=r, column=2, value=str(v))
        ws3.column_dimensions["A"].width = 24
        ws3.column_dimensions["B"].width = 32

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = (
            f'attachment; filename="{group.code}_report_{label}.xlsx"')
        wb.save(resp)
        return resp
