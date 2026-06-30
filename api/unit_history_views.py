from __future__ import annotations

from datetime import date, timedelta
from typing import Dict, Iterable, List, Optional, Tuple

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from api.active_day import get_active_date
from api.models import (
    OperationType,
    Schedule,
    Task,
    TaskStatus,
    Technician,
    TechnicianRole,
    Unit,
    UnitMaintenanceState,
)

PAGE_SIZE_MAX = 200
DUE_SOON_DAYS = 7
REPEAT_CALLBACK_WARNING = 3
REPEAT_CALLBACK_CRITICAL = 5


def _group(request):
    return getattr(request.user, "supervised_group", None)


def _group_type(group) -> str:
    roles = set(
        Technician.objects.filter(group=group, is_active_employee=True)
        .values_list("tech_role", flat=True)
    )
    if roles and roles.issubset({TechnicianRole.CALLBACK}):
        return OperationType.CALLBACK
    return OperationType.MAINTENANCE


def _parse_as_of(request) -> Optional[date]:
    raw = (request.query_params.get("as_of") or "").strip()
    if raw.lower() == "all":
        return None
    if raw:
        try:
            return date.fromisoformat(raw[:10])
        except ValueError:
            pass
    return get_active_date()


def _as_of_date_for_metrics(group, as_of: Optional[date]) -> date:
    if as_of is not None:
        return as_of
    last = (
        Schedule.objects.filter(technician__group=group)
        .order_by("-start_time")
        .values_list("start_time", flat=True)
        .first()
    )
    if last:
        return last.date()
    return get_active_date()


def _schedule_scope(group, as_of: Optional[date], operation: str):
    qs = (
        Schedule.objects.filter(
            technician__group=group,
            task__task_type__operation_type=operation,
            start_time__isnull=False,
        )
        .select_related("technician", "task", "task__unit", "task__task_type")
        .order_by("start_time", "sequence_order")
    )
    if as_of is not None:
        qs = qs.filter(start_time__date__lte=as_of)
    return qs


def _unassigned_scope(group, as_of: Optional[date], operation: str):
    qs = Task.objects.filter(
        is_active=True,
        is_unassigned=True,
        task_type__operation_type=operation,
    ).select_related("unit", "task_type", "assigned_group")

    # If the task has an assigned group, respect it. Legacy simulated callback
    # backlog may have assigned_group=NULL; keep those visible in callback HQ so
    # the supervisor can see the real backlog instead of losing it.
    qs = qs.filter(Q(assigned_group=group) | Q(assigned_group__isnull=True))
    if as_of is not None:
        qs = qs.filter(release_time__date__lte=as_of)
    return qs.order_by("release_time", "priority", "unit__unit_name")


def _minutes(start, end) -> int:
    if not start or not end:
        return 0
    return max(0, int((end - start).total_seconds() // 60))


def _sla_met(task: Task, start_time) -> Optional[bool]:
    if not task or task.task_type.operation_type != OperationType.CALLBACK:
        return None
    target = task.task_type.sla_target_min
    if not target or not task.release_time or not start_time:
        return None
    delta = int((start_time - task.release_time).total_seconds() // 60)
    return delta <= int(target)


def _sla_minutes(task: Task, start_time) -> Optional[int]:
    if not task or task.task_type.operation_type != OperationType.CALLBACK:
        return None
    if not task.release_time or not start_time:
        return None
    return max(0, int((start_time - task.release_time).total_seconds() // 60))


def _last_service(schedules: Iterable[Schedule]) -> Optional[date]:
    vals = [s.start_time.date() for s in schedules if s.start_time]
    return max(vals) if vals else None


def _next_planned(group, unit_id: int, as_of_day: date, operation: str) -> Optional[date]:
    q = (
        Schedule.objects.filter(
            technician__group=group,
            task__unit_id=unit_id,
            task__task_type__operation_type=operation,
            start_time__date__gt=as_of_day,
        )
        .order_by("start_time")
        .values_list("start_time", flat=True)
        .first()
    )
    return q.date() if q else None


def _maintenance_due(unit: Unit, as_of_day: date) -> Tuple[str, Optional[date], Optional[int]]:
    try:
        state = unit.maintenance_state
    except UnitMaintenanceState.DoesNotExist:
        return "OVERDUE", None, None

    due_type = state.due_type(as_of_day)
    if due_type:
        return "OVERDUE", None, None

    candidates = []
    if state.last_c_date:
        candidates.append(state.last_c_date + timedelta(days=UnitMaintenanceState.C_DAYS))
    if state.last_b_date:
        candidates.append(state.last_b_date + timedelta(days=UnitMaintenanceState.B_DAYS))
    if state.last_a_date:
        candidates.append(state.last_a_date + timedelta(days=UnitMaintenanceState.A_DAYS))
    if not candidates:
        return "OVERDUE", None, None
    nxt = min(candidates)
    days = (nxt - as_of_day).days
    if days <= DUE_SOON_DAYS:
        return "DUE_SOON", nxt, days
    return "HEALTHY", nxt, days


def _row_status(operation: str, unit: Unit, as_of_day: date, data: Dict) -> Tuple[str, str]:
    unassigned = int(data.get("unassigned_count", 0) or 0)
    aa_unassigned = int(data.get("aa_unassigned", 0) or 0)
    repeat_callbacks = int(data.get("repeat_callbacks", 0) or 0)
    sla_pct = data.get("sla_pct")

    if operation == OperationType.CALLBACK:
        if aa_unassigned > 0:
            return "critical", "AA backlog"
        if unassigned > 0:
            return "unassigned", "Callback backlog"
        if sla_pct is not None and float(sla_pct) < 80 and data.get("callback", 0):
            return "sla_risk", "SLA risk"
        if repeat_callbacks >= REPEAT_CALLBACK_CRITICAL:
            return "callback_risk", "Repeat critical"
        if repeat_callbacks >= REPEAT_CALLBACK_WARNING:
            return "callback_risk", "Repeat callback"
        return "healthy", "Healthy"

    # Maintenance
    if unassigned > 0:
        return "unassigned", "Maintenance backlog"
    due_status, _nxt, _days = _maintenance_due(unit, as_of_day)
    if due_status == "OVERDUE":
        return "overdue", "Overdue"
    if due_status == "DUE_SOON":
        return "due_soon", "Due soon"
    return "healthy", "Healthy"


def _build_unit_rows(group, operation: str, as_of: Optional[date]) -> Tuple[List[Dict], Dict]:
    as_of_day = _as_of_date_for_metrics(group, as_of)
    schedules = list(_schedule_scope(group, as_of, operation))
    unassigned = list(_unassigned_scope(group, as_of, operation))

    by_unit: Dict[int, Dict] = {}
    units_by_id: Dict[int, Unit] = {}

    for s in schedules:
        u = s.task.unit
        units_by_id[u.id] = u
        row = by_unit.setdefault(u.id, {
            "unit": u,
            "maint": 0,
            "callback": 0,
            "aa_count": 0,
            "b_count": 0,
            "sla_total": 0,
            "sla_met": 0,
            "last": None,
            "unassigned_count": 0,
            "aa_unassigned": 0,
            "b_unassigned": 0,
        })
        op = s.task.task_type.operation_type
        if op == OperationType.MAINTENANCE:
            row["maint"] += 1
        elif op == OperationType.CALLBACK:
            row["callback"] += 1
            if s.task.priority == "AA":
                row["aa_count"] += 1
            elif s.task.priority == "B":
                row["b_count"] += 1
            met = _sla_met(s.task, s.start_time)
            if met is not None:
                row["sla_total"] += 1
                if met:
                    row["sla_met"] += 1
        d = s.start_time.date()
        if row["last"] is None or d > row["last"]:
            row["last"] = d

    for t in unassigned:
        u = t.unit
        units_by_id[u.id] = u
        row = by_unit.setdefault(u.id, {
            "unit": u,
            "maint": 0,
            "callback": 0,
            "aa_count": 0,
            "b_count": 0,
            "sla_total": 0,
            "sla_met": 0,
            "last": None,
            "unassigned_count": 0,
            "aa_unassigned": 0,
            "b_unassigned": 0,
        })
        row["unassigned_count"] += 1
        if t.priority == "AA":
            row["aa_unassigned"] += 1
        elif t.priority == "B":
            row["b_unassigned"] += 1

    # Repeat callback window is independent of the current group role because it
    # is a unit-health signal. Use the 30 days ending at the roll-date.
    window_start = as_of_day - timedelta(days=30)
    repeat_counts = {}
    cb_qs = Schedule.objects.filter(
        task__unit_id__in=units_by_id.keys(),
        task__task_type__operation_type=OperationType.CALLBACK,
        start_time__date__gte=window_start,
        start_time__date__lte=as_of_day,
    ).values_list("task__unit_id", flat=True)
    for uid in cb_qs:
        repeat_counts[uid] = repeat_counts.get(uid, 0) + 1

    result = []
    for uid, row in by_unit.items():
        u = row["unit"]
        due_status, due_date, days_until_due = _maintenance_due(u, as_of_day)
        next_planned = _next_planned(group, uid, as_of_day, operation)
        repeat_callbacks = repeat_counts.get(uid, 0)
        row["repeat_callbacks"] = repeat_callbacks
        if row["sla_total"]:
            row["sla_pct"] = round(row["sla_met"] * 100.0 / row["sla_total"], 1)
        else:
            row["sla_pct"] = None
        status, status_label = _row_status(operation, u, as_of_day, row)
        result.append({
            "id": uid,
            "name": u.unit_name,
            "code": u.unit_code,
            "district": u.district or "",
            "type": u.unit_type,
            "maint": row["maint"],
            "callback": row["callback"],
            "aa_count": row["aa_count"],
            "b_count": row["b_count"],
            "sla_pct": row["sla_pct"],
            "last": row["last"].isoformat() if row["last"] else None,
            "next_service": next_planned.isoformat() if next_planned else (due_date.isoformat() if due_date else None),
            "days_until_due": days_until_due,
            "due_status": due_status,
            "repeat_callbacks": repeat_callbacks,
            "unassigned_count": row["unassigned_count"],
            "aa_unassigned": row["aa_unassigned"],
            "b_unassigned": row["b_unassigned"],
            "status": status,
            "status_label": status_label,
        })

    summary = {
        "units": len(result),
        "serviced_units": sum(1 for r in result if r["last"]),
        "due_soon": sum(1 for r in result if r["status"] == "due_soon"),
        "overdue": sum(1 for r in result if r["status"] == "overdue"),
        "callback_risk": sum(1 for r in result if r["status"] == "callback_risk"),
        "sla_risk": sum(1 for r in result if r["status"] == "sla_risk"),
        "unassigned": sum(r["unassigned_count"] for r in result),
        "aa_unassigned": sum(r["aa_unassigned"] for r in result),
        "b_unassigned": sum(r["b_unassigned"] for r in result),
        "aa_count": sum(r["aa_count"] for r in result),
        "b_count": sum(r["b_count"] for r in result),
        "repeat_callback_units": sum(1 for r in result if r["repeat_callbacks"] >= REPEAT_CALLBACK_WARNING),
    }
    return result, summary


def _sort_rows(rows: List[Dict], sort: str, order: str) -> List[Dict]:
    reverse = order != "asc"
    if sort == "name":
        key = lambda r: (r["name"] or "").lower()
        reverse = order == "desc"
    elif sort == "maint":
        key = lambda r: r["maint"]
    elif sort == "callback":
        key = lambda r: r["callback"]
    elif sort == "next_service":
        key = lambda r: r["next_service"] or "9999-99-99"
        reverse = order == "desc"
    elif sort == "status":
        priority = {"critical": 0, "unassigned": 1, "overdue": 2, "sla_risk": 3, "callback_risk": 4, "due_soon": 5, "healthy": 6}
        key = lambda r: priority.get(r["status"], 99)
        reverse = False if order == "asc" else True
    elif sort == "unassigned":
        key = lambda r: r["unassigned_count"]
    elif sort == "sla":
        key = lambda r: r["sla_pct"] if r["sla_pct"] is not None else -1
    else:
        key = lambda r: r["last"] or "0000-00-00"
    return sorted(rows, key=key, reverse=reverse)


class UnitHistorySummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group = _group(request)
        if not group:
            return Response({"detail": "No supervisor group for this user."}, status=403)
        operation = _group_type(group)
        as_of = _parse_as_of(request)
        rows, summary = _build_unit_rows(group, operation, as_of)

        search = (request.query_params.get("search") or "").strip().lower()
        status_filter = (request.query_params.get("status") or "all").strip().lower()
        if search:
            rows = [r for r in rows if search in (r["name"] or "").lower() or search in (r["code"] or "").lower()]
        if status_filter and status_filter != "all":
            if status_filter == "risk":
                rows = [r for r in rows if r["status"] in {"critical", "unassigned", "overdue", "sla_risk", "callback_risk"}]
            else:
                rows = [r for r in rows if r["status"] == status_filter]

        sort = request.query_params.get("sort") or "last_service"
        order = request.query_params.get("order") or "desc"
        rows = _sort_rows(rows, sort, order)

        try:
            page = max(1, int(request.query_params.get("page") or 1))
        except ValueError:
            page = 1
        try:
            page_size = min(PAGE_SIZE_MAX, max(1, int(request.query_params.get("page_size") or 50)))
        except ValueError:
            page_size = 50
        total = len(rows)
        start = (page - 1) * page_size
        end = start + page_size

        return Response({
            "group_type": operation,
            "scope": "full" if as_of is None else "roll_date",
            "as_of": as_of.isoformat() if as_of else "all",
            "summary": summary,
            "total_units": total,
            "units": rows[start:end],
        })


class UnitHistoryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, unit_id: int):
        group = _group(request)
        if not group:
            return Response({"detail": "No supervisor group for this user."}, status=403)
        operation = _group_type(group)
        as_of = _parse_as_of(request)
        unit = Unit.objects.get(id=unit_id)

        schedules = _schedule_scope(group, as_of, operation).filter(task__unit_id=unit_id)
        visits = []
        for s in schedules:
            task = s.task
            is_cb = task.task_type.operation_type == OperationType.CALLBACK
            sla = _sla_met(task, s.start_time)
            response = _sla_minutes(task, s.start_time)
            visits.append({
                "date": s.start_time.date().isoformat(),
                "kind": "Callback" if is_cb else "Maintenance",
                "operation": task.task_type.operation_type,
                "task_no": task.task_no,
                "type": task.priority if is_cb else (task.task_type.maintenance_type or task.task_type.code),
                "start": s.start_time.strftime("%H:%M"),
                "end": s.end_time.strftime("%H:%M") if s.end_time else None,
                "technician": s.technician.full_name,
                "duration_min": _minutes(s.start_time, s.end_time),
                "travel_min": s.travel_time_min or 0,
                "route_km": float(s.travel_distance_km or 0),
                "sla_met": sla,
                "response_min": response,
                "release_time": task.release_time.strftime("%Y-%m-%d %H:%M") if task.release_time else None,
            })

        unassigned = []
        for t in _unassigned_scope(group, as_of, operation).filter(unit_id=unit_id):
            unassigned.append({
                "task_no": t.task_no,
                "operation": t.task_type.operation_type,
                "type": t.priority if t.task_type.operation_type == OperationType.CALLBACK else (t.task_type.maintenance_type or t.task_type.code),
                "release_time": t.release_time.strftime("%Y-%m-%d %H:%M") if t.release_time else None,
                "duration_min": t.estimated_duration_min,
                "reason": t.unassigned_reason or "Unassigned",
            })

        return Response({
            "unit": {"id": unit.id, "name": unit.unit_name, "code": unit.unit_code, "type": unit.unit_type},
            "visits": visits,
            "unassigned": unassigned,
        })


class UnitHistoryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group = _group(request)
        if not group:
            return Response({"detail": "No supervisor group for this user."}, status=403)
        operation = _group_type(group)
        as_of = _parse_as_of(request)
        rows, summary = _build_unit_rows(group, operation, as_of)
        search = (request.query_params.get("search") or "").strip().lower()
        status_filter = (request.query_params.get("status") or "all").strip().lower()
        if search:
            rows = [r for r in rows if search in (r["name"] or "").lower() or search in (r["code"] or "").lower()]
        if status_filter and status_filter != "all":
            if status_filter == "risk":
                rows = [r for r in rows if r["status"] in {"critical", "unassigned", "overdue", "sla_risk", "callback_risk"}]
            else:
                rows = [r for r in rows if r["status"] == status_filter]
        rows = _sort_rows(rows, request.query_params.get("sort") or "last_service", request.query_params.get("order") or "desc")

        wb = Workbook()
        ws = wb.active
        ws.title = "Unit Summary"
        headers = ["Unit", "Code", "Type", "Maintenance", "Callbacks", "AA", "B", "SLA %", "Last Service", "Next Service", "Due Days", "Repeat CB 30d", "Unassigned", "AA Backlog", "B Backlog", "Status"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for r in rows:
            ws.append([
                r["name"], r["code"], r["type"], r["maint"], r["callback"],
                r["aa_count"], r["b_count"], r["sla_pct"] if r["sla_pct"] is not None else "N/A",
                r["last"] or "", r["next_service"] or "", r["days_until_due"] if r["days_until_due"] is not None else "",
                r["repeat_callbacks"], r["unassigned_count"], r["aa_unassigned"], r["b_unassigned"], r["status_label"],
            ])

        ws2 = wb.create_sheet("Visit Detail")
        detail_headers = ["Unit", "Code", "Date", "Task", "Operation", "Type/Priority", "Start", "End", "Technician", "Duration", "Travel Min", "Route KM", "Response Min", "SLA Met"]
        ws2.append(detail_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        schedules = _schedule_scope(group, as_of, operation)
        for s in schedules:
            task = s.task
            is_cb = task.task_type.operation_type == OperationType.CALLBACK
            sla = _sla_met(task, s.start_time)
            ws2.append([
                task.unit.unit_name, task.unit.unit_code, s.start_time.date().isoformat(), task.task_no,
                task.task_type.operation_type, task.priority if is_cb else (task.task_type.maintenance_type or task.task_type.code),
                s.start_time.strftime("%H:%M"), s.end_time.strftime("%H:%M") if s.end_time else "",
                s.technician.full_name, _minutes(s.start_time, s.end_time), s.travel_time_min or 0,
                float(s.travel_distance_km or 0), _sla_minutes(task, s.start_time) or "", "YES" if sla else ("NO" if sla is False else "N/A"),
            ])

        ws3 = wb.create_sheet("Unassigned Backlog")
        ws3.append(["Task", "Unit", "Code", "Operation", "Priority/Type", "Release Time", "Duration", "Reason"])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
        for t in _unassigned_scope(group, as_of, operation):
            ws3.append([
                t.task_no, t.unit.unit_name, t.unit.unit_code, t.task_type.operation_type,
                t.priority or t.task_type.maintenance_type or t.task_type.code,
                t.release_time.strftime("%Y-%m-%d %H:%M") if t.release_time else "",
                t.estimated_duration_min, t.unassigned_reason or "Unassigned",
            ])

        ws4 = wb.create_sheet("Metrics")
        for k, v in summary.items():
            ws4.append([k, v])
        ws4.append(["operation", operation])
        ws4.append(["as_of", as_of.isoformat() if as_of else "all"])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="unit_history_enhanced.xlsx"'
        wb.save(response)
        return response
