from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from api.models import (
    SpecialtyType,
    SupervisorGroup,
    Technician,
    TechnicianRole,
    UserProfile,
    UserRole,
)


ROLE_MAP = {
    "bakımcı": TechnicianRole.MAINTENANCE,
    "bakimci": TechnicianRole.MAINTENANCE,
    "maintenance": TechnicianRole.MAINTENANCE,
    "arızacı": TechnicianRole.CALLBACK,
    "arizaci": TechnicianRole.CALLBACK,
    "callback": TechnicianRole.CALLBACK,
}

SPECIALTY_MAP = {
    "asansör": SpecialtyType.ELEVATOR,
    "asansor": SpecialtyType.ELEVATOR,
    "elevator": SpecialtyType.ELEVATOR,
    "yürüyen merdiven": SpecialtyType.ESCALATOR,
    "yuruyen merdiven": SpecialtyType.ESCALATOR,
    "escalator": SpecialtyType.ESCALATOR,
    "asansör + yürüyen merdiven": SpecialtyType.BOTH,
    "asansor + yuruyen merdiven": SpecialtyType.BOTH,
    "asansör+yürüyen merdiven": SpecialtyType.BOTH,
    "asansor+yuruyen merdiven": SpecialtyType.BOTH,
    "both": SpecialtyType.BOTH,
}


def norm(value: Any) -> str:
    return str(value or "").strip()


def key(value: Any) -> str:
    text = norm(value).lower()
    return (
        text.replace("ı", "i")
        .replace("İ", "i")
        .replace("ö", "o")
        .replace("ü", "u")
        .replace("ğ", "g")
        .replace("ş", "s")
        .replace("ç", "c")
    )


def parse_role(value: Any) -> str:
    k = key(value)
    if k in ROLE_MAP:
        return ROLE_MAP[k]
    raise CommandError(f"Unknown Görev Türü / role value: {value!r}")


def parse_specialty(value: Any) -> str:
    raw = norm(value)
    k = key(raw)
    if "+" in raw or "both" in k:
        return SpecialtyType.BOTH
    if k in SPECIALTY_MAP:
        return SPECIALTY_MAP[k]
    # Safe default used by the project seed data.
    return SpecialtyType.BOTH


def find_group(spv_name: str) -> SupervisorGroup:
    spv = norm(spv_name)
    if not spv:
        raise CommandError("SPV column is empty for a roster row.")

    candidates = [
        f"{spv} Group",
        spv,
    ]
    for name in candidates:
        group = SupervisorGroup.objects.filter(name=name).first()
        if group:
            return group

    group = SupervisorGroup.objects.filter(name__icontains=spv).first()
    if group:
        return group

    raise CommandError(f"Could not find SupervisorGroup for SPV={spv!r}")


def row_to_dict(headers: list[str], row: tuple[Any, ...]) -> Dict[str, Any]:
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


class Command(BaseCommand):
    help = "Apply the rightsized technician roster Excel to the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            help="Path to the roster Excel file, e.g. data/technician_roster_final.xlsx",
        )
        parser.add_argument(
            "--sheet",
            default="Teknisyenler",
            help="Active roster sheet name. Default: Teknisyenler",
        )
        parser.add_argument(
            "--password",
            default="tech12345",
            help="Password to set when --reset-password is used. Default: tech12345",
        )
        parser.add_argument(
            "--reset-password",
            action="store_true",
            help="Reset active technician user passwords to --password.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print what would change without saving.",
        )

    def handle(self, *args, **options):
        path = Path(options["excel_path"])
        if not path.exists():
            raise CommandError(f"Excel file not found: {path}")

        wb = load_workbook(path, data_only=True)
        sheet_name = options["sheet"]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        required = ["Technician ID", "Ad", "Soyad", "Görev Türü", "Uzmanlık", "Yetkinlik Kodu", "SPV"]
        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Missing required Excel columns: {missing}. Found: {headers}")

        active_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = row_to_dict(headers, row)
            employee_code = norm(data.get("Technician ID"))
            if not employee_code:
                continue
            active_rows.append(data)

        active_codes = {norm(r["Technician ID"]) for r in active_rows}
        if not active_codes:
            raise CommandError("No active technicians found in Excel sheet.")

        self.stdout.write(self.style.NOTICE(f"Excel active roster rows: {len(active_codes)}"))

        dry_run = options["dry_run"]
        reset_password = options["reset_password"]
        password = options["password"]

        updated = 0
        created = 0
        reactivated = 0
        deactivated = 0

        def apply_changes():
            nonlocal updated, created, reactivated, deactivated

            # Deactivate every technician not listed in the Excel active roster.
            to_deactivate = Technician.objects.exclude(employee_code__in=active_codes).filter(is_active_employee=True)
            deactivated = to_deactivate.count()
            if not dry_run:
                to_deactivate.update(is_active_employee=False, last_location_at=None)

            for data in active_rows:
                code = norm(data["Technician ID"])
                first = norm(data["Ad"])
                last = norm(data["Soyad"])
                full_name = f"{first} {last}".strip()
                group = find_group(norm(data["SPV"]))
                tech_role = parse_role(data["Görev Türü"])
                specialty = parse_specialty(data["Uzmanlık"])
                competency_code = norm(data.get("Yetkinlik Kodu")) or None
                username = code.lower()

                tech = Technician.objects.filter(employee_code=code).select_related("user").first()

                if tech is None:
                    user = User.objects.filter(username=username).first()
                    if user is None:
                        user = User(username=username, first_name=first, last_name=last)
                        user.set_password(password)
                        if not dry_run:
                            user.save()
                    elif reset_password:
                        user.set_password(password)
                        if not dry_run:
                            user.save()

                    if not dry_run:
                        UserProfile.objects.get_or_create(user=user, defaults={"role": UserRole.TECH})
                        tech = Technician.objects.create(
                            user=user,
                            employee_code=code,
                            competency_code=competency_code,
                            full_name=full_name,
                            group=group,
                            tech_role=tech_role,
                            specialty=specialty,
                            is_active_employee=True,
                        )
                    created += 1
                    continue

                was_inactive = not tech.is_active_employee
                if was_inactive:
                    reactivated += 1

                tech.full_name = full_name
                tech.group = group
                tech.tech_role = tech_role
                tech.specialty = specialty
                tech.competency_code = competency_code
                tech.is_active_employee = True
                tech.last_location_at = None

                user = tech.user
                user.first_name = first
                user.last_name = last
                if reset_password:
                    user.set_password(password)

                if not dry_run:
                    user.save()
                    UserProfile.objects.update_or_create(user=user, defaults={"role": UserRole.TECH})
                    tech.save()
                updated += 1

        if dry_run:
            apply_changes()
        else:
            with transaction.atomic():
                apply_changes()

        self.stdout.write(self.style.SUCCESS("Roster apply complete." if not dry_run else "Dry run complete."))
        self.stdout.write(f"Active roster target: {len(active_codes)}")
        self.stdout.write(f"Updated active technicians: {updated}")
        self.stdout.write(f"Created missing technicians: {created}")
        self.stdout.write(f"Reactivated technicians: {reactivated}")
        self.stdout.write(f"Deactivated technicians not in Excel: {deactivated}")

        self.stdout.write("\nActive counts by supervisor group:")
        for group in SupervisorGroup.objects.all().order_by("name"):
            qs = Technician.objects.filter(group=group, is_active_employee=True)
            total = qs.count()
            if total:
                maintenance = qs.filter(tech_role=TechnicianRole.MAINTENANCE).count()
                callback = qs.filter(tech_role=TechnicianRole.CALLBACK).count()
                self.stdout.write(f"- {group.name}: active={total}, maintenance={maintenance}, callback={callback}")
